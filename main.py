"""
DeepFRI Web API Script for Windows
Uses the DeepFRI web service to predict protein functions
Much simpler than local installation - no compilation issues!
"""

import os
import time
import requests
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from pathlib import Path
import glob
from tqdm import tqdm

class DeepFRIWebAPI:
    """Process PDB files using DeepFRI web API."""
    
    def __init__(self, pdb_folder, output_excel="protein_function_results.xlsx"):
        self.pdb_folder = pdb_folder
        self.output_excel = output_excel
        # DeepFRI beta web service endpoint
        self.base_url = "https://beta.deepfri.flatironinstitute.org"
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'DeepFRI-Python-Script/1.0'
        })
    
    def upload_pdb(self, pdb_file_path):
        """
        Upload a PDB file to DeepFRI web service.
        Returns job ID if successful.
        """
        try:
            with open(pdb_file_path, 'r') as f:
                pdb_content = f.read()
            
            # Upload endpoint (you may need to adjust this based on actual API)
            upload_url = f"{self.base_url}/api/predict"
            
            files = {
                'pdb_file': (os.path.basename(pdb_file_path), pdb_content, 'chemical/x-pdb')
            }
            
            data = {
                'job_name': os.path.splitext(os.path.basename(pdb_file_path))[0]
            }
            
            response = self.session.post(upload_url, files=files, data=data, timeout=30)
            
            if response.status_code == 200:
                result = response.json()
                return result.get('job_id') or result.get('id')
            else:
                print(f"Upload failed with status {response.status_code}")
                return None
                
        except Exception as e:
            print(f"Error uploading {pdb_file_path}: {str(e)}")
            return None
    
    def check_job_status(self, job_id, max_wait=300):
        """
        Check if job is complete and return results.
        Polls every 5 seconds up to max_wait seconds.
        """
        status_url = f"{self.base_url}/api/status/{job_id}"
        elapsed = 0
        
        while elapsed < max_wait:
            try:
                response = self.session.get(status_url, timeout=10)
                if response.status_code == 200:
                    status_data = response.json()
                    
                    if status_data.get('status') == 'complete':
                        return status_data
                    elif status_data.get('status') == 'failed':
                        return None
                
                time.sleep(5)
                elapsed += 5
                
            except Exception as e:
                print(f"Error checking status: {str(e)}")
                time.sleep(5)
                elapsed += 5
        
        print(f"Job {job_id} timed out after {max_wait}s")
        return None
    
    def parse_predictions(self, results_data):
        """
        Parse prediction results from API response.
        Returns structured data for each ontology.
        """
        parsed = {
            'molecular_function': [],
            'biological_process': [],
            'enzyme_commission': [],
            'used_sequence_based': {
                'molecular_function': False,
                'biological_process': False,
                'enzyme_commission': False
            }
        }
        
        # Map API response keys to our categories
        ontology_map = {
            'molecular_function': ['mf', 'molecular_function', 'MF'],
            'biological_process': ['bp', 'biological_process', 'BP'],
            'enzyme_commission': ['ec', 'enzyme_commission', 'EC']
        }
        
        for category, api_keys in ontology_map.items():
            # Try structure-based first
            structure_found = False
            for key in api_keys:
                structure_key = f'structure_{key}'
                if structure_key in results_data:
                    predictions = results_data[structure_key]
                    if predictions and len(predictions) > 0:
                        parsed[category] = self._format_predictions(predictions[:3])
                        structure_found = True
                        break
            
            # Fall back to sequence-based if no structure predictions
            if not structure_found:
                for key in api_keys:
                    sequence_key = f'sequence_{key}'
                    if sequence_key in results_data:
                        predictions = results_data[sequence_key]
                        if predictions and len(predictions) > 0:
                            parsed[category] = self._format_predictions(predictions[:3])
                            parsed['used_sequence_based'][category] = True
                            break
        
        return parsed
    
    def _format_predictions(self, predictions):
        """Format predictions as list of strings."""
        formatted = []
        for pred in predictions:
            # Expected format: {"term": "GO:XXXXX", "name": "function name", "score": 0.99}
            term = pred.get('term', pred.get('go_term', pred.get('id', '')))
            name = pred.get('name', pred.get('description', ''))
            score = pred.get('score', pred.get('confidence', 0))
            
            formatted.append(f"{name} {term} {score:.2f}")
        
        return formatted
    
    def process_pdb_file(self, pdb_file_path):
        """Process a single PDB file through the API."""
        protein_id = os.path.splitext(os.path.basename(pdb_file_path))[0]
        
        print(f"  Uploading {protein_id}...", end=" ")
        job_id = self.upload_pdb(pdb_file_path)
        
        if not job_id:
            print("❌ Upload failed")
            return None
        
        print(f"✓ (Job ID: {job_id})")
        print(f"  Waiting for results...", end=" ")
        
        results = self.check_job_status(job_id)
        
        if not results:
            print("❌ Processing failed")
            return None
        
        print("✓")
        
        parsed_results = self.parse_predictions(results)
        parsed_results['protein_id'] = protein_id
        
        return parsed_results
    
    def process_all_pdbs(self):
        """Process all PDB files in the folder."""
        pdb_files = sorted(glob.glob(os.path.join(self.pdb_folder, "*.pdb")))
        
        if not pdb_files:
            print(f"\n❌ No PDB files found in: {self.pdb_folder}")
            return []
        
        print(f"\n📁 Found {len(pdb_files)} PDB files")
        print("=" * 70)
        
        all_results = []
        
        for idx, pdb_file in enumerate(pdb_files, 1):
            print(f"\n[{idx}/{len(pdb_files)}]")
            result = self.process_pdb_file(pdb_file)
            
            if result:
                all_results.append(result)
                print("  ✅ Complete")
            else:
                print("  ⚠️  Failed - skipping")
            
            # Be respectful to the server
            if idx < len(pdb_files):
                time.sleep(2)
        
        return all_results
    
    def create_excel_output(self, results_list):
        """Create formatted Excel file with results."""
        if not results_list:
            print("\n❌ No results to save")
            return
        
        print("\n" + "=" * 70)
        print("Creating Excel file...")
        
        # Prepare data
        data = []
        highlight_info = []
        
        for result in results_list:
            row = {
                'Protein ID': result['protein_id'],
                'Molecular Function': '\n'.join(result['molecular_function']) if result['molecular_function'] else '',
                'Biological Process': '\n'.join(result['biological_process']) if result['biological_process'] else '',
                'Enzyme Commission': '\n'.join(result['enzyme_commission']) if result['enzyme_commission'] else '',
            }
            data.append(row)
            highlight_info.append(result['used_sequence_based'])
        
        # Create DataFrame and save
        df = pd.DataFrame(data)
        df.to_excel(self.output_excel, index=False, engine='openpyxl')
        
        # Apply formatting
        wb = load_workbook(self.output_excel)
        ws = wb.active
        
        # Define styles
        light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        header_font = Font(bold=True, size=11)
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply cell formatting
        for idx, used_seq in enumerate(highlight_info, start=2):
            # Molecular Function
            ws[f'B{idx}'].alignment = wrap_alignment
            if used_seq['molecular_function']:
                ws[f'B{idx}'].fill = light_red_fill
            
            # Biological Process
            ws[f'C{idx}'].alignment = wrap_alignment
            if used_seq['biological_process']:
                ws[f'C{idx}'].fill = light_red_fill
            
            # Enzyme Commission
            ws[f'D{idx}'].alignment = wrap_alignment
            if used_seq['enzyme_commission']:
                ws[f'D{idx}'].fill = light_red_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 60
        
        # Adjust row heights
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 60
        
        wb.save(self.output_excel)
        print(f"✅ Results saved to: {self.output_excel}")
        print("=" * 70)
    
    def run(self):
        """Main execution method."""
        print("=" * 70)
        print("     DeepFRI Protein Function Prediction (Web API)")
        print("=" * 70)
        print(f"Input folder: {os.path.abspath(self.pdb_folder)}")
        print(f"Output file:  {self.output_excel}")
        
        results = self.process_all_pdbs()
        
        if results:
            self.create_excel_output(results)
            print(f"\n🎉 Successfully processed {len(results)} proteins!")
        else:
            print("\n⚠️  No proteins were processed successfully")


# ============================================================================
#                           CONFIGURATION
# ============================================================================

if __name__ == "__main__":
    # IMPORTANT: Update this path to your PDB folder
    
    PDB_FOLDER = "D:/Work/Open Source/protein functions/pdb_files"
    OUTPUT_FILE = "results.xlsx"
    
    # ========================================================================
    
    # Validate folder exists
    if not os.path.exists(PDB_FOLDER):
        print("\n" + "=" * 70)
        print("❌ ERROR: Folder not found!")
        print("=" * 70)
        print(f"\nThe folder '{PDB_FOLDER}' does not exist.")
        print("\n📝 To fix this:")
        print("1. Open this script in Notepad")
        print("2. Find the line: PDB_FOLDER = r\"C:\\path\\to\\your\\pdb\\folder\"")
        print("3. Change it to your actual folder path")
        print("\n💡 Example paths:")
        print("   PDB_FOLDER = r\"C:\\Users\\YourName\\Documents\\PDB_Files\"")
        print("   PDB_FOLDER = r\"D:\\Research\\Proteins\"")
        print("=" * 70)
    else:
        # Run the processor
        processor = DeepFRIWebAPI(PDB_FOLDER, OUTPUT_FILE)
        
        print("\n⚠️  IMPORTANT NOTE:")
        print("This script uses the DeepFRI web API. The actual API endpoints")
        print("may need to be adjusted based on the current API documentation.")
        print("If this doesn't work, you may need to inspect the web interface")
        print("and update the API URLs in the script.")
        print("\n" + "=" * 70)
        
        input("\nPress Enter to continue...")
        
        processor.run()
        
        print("\n\nPress Enter to exit...")
        input()